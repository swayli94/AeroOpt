'''
Database definition.
'''

import numpy as np
import json
import copy
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from typing import List, Tuple, Callable
from AeroOpt.core.individual import Individual
from AeroOpt.core.problem import Problem
from AeroOpt.core.mpEvaluation import MultiProcessEvaluation


class Database(object):
    '''
    Basic database class.
    
    Parameters:
    -----------
    problem: Problem
        Problem of the database.
    database_type: str
        Type of the database.
    '''
    
    database_type_dict = {
        'default':      'not specified',
        'elite':        'elite database',
        'valid':        'valid database',
        'total':        'total database',
        'population':   'population database',
        'surrogate':    'surrogate model',
        'diversity':    'diversity calculation',
        'intersection': 'intersection of two databases',
        'sub-database': 'sub-database',
    }

    def __init__(self, problem: Problem, database_type='default'):

        self.problem = problem

        if database_type not in self.database_type_dict:
            raise ValueError(f'Invalid database type: {database_type}')
        self.database_type = database_type
        
        self.individuals : List[Individual] = []
        
        self._id_list : List[int] = [] # List of IDs in the order of individuals
        self._index_pareto_fronts : List[List[int]] = [] # Pareto fronts of individual indices
        
        self._sorted : bool = False
        self._updated_crowding_distance : bool = False
        self._updated_pareto_rank : bool = False
        self._is_valid_database : bool = False

    #* Attributes
    
    @property
    def size(self) -> int:
        '''
        Size of the database.
        '''
        return len(self.individuals)
    
    @property
    def sorted(self) -> bool:
        '''
        Whether the database is sorted.
        '''
        return self._sorted
    
    @property
    def is_valid_database(self) -> bool:
        '''
        Whether the database only contains valid individuals.
        '''
        return self._is_valid_database
    
    @property
    def updated_crowding_distance(self) -> bool:
        '''
        Whether the crowding distance is updated.
        '''
        return self._updated_crowding_distance
    
    @property
    def updated_pareto_rank(self) -> bool:
        '''
        Whether the Pareto rank is updated.
        '''
        return self._updated_pareto_rank
    
    def critical_scaled_distance(self) -> float:
        '''
        Critical scaled distance for checking duplication of individuals.
        '''
        return self.problem.data_settings.critical_scaled_distance
        
    #* Basic functions
    
    def empty_database(self) -> None:
        '''
        Empty the database.
        '''
        self.individuals = []
        self._id_list = []
        self._sorted = False
        self._updated_crowding_distance = False
        self._updated_pareto_rank = False
        self._is_valid_database = False
    
    def copy_from_database(self, other: 'Database', 
                        ID_list: List[int] = None, 
                        index_list: List[int] = None,
                        deepcopy: bool = True) -> None:
        '''
        Copy the database from another database:
        
        - If `ID_list` or `index_list` is provided, the specified individuals are copied.
        - Otherwise, the entire database is copied.
        - The `database_type` is not changed.
        
        Parameters:
        -----------
        other: Database
            Another database to copy from.
        ID_list: List[int]
            List of IDs of individuals to be selected. This has higher priority than `index_list`.
        index_list: List[int]
            List of index of individuals to be selected.
        deepcopy: bool
            If True, the individuals are copied.
        '''
        if other.problem is not self.problem:
            raise ValueError('Databases must share the same problem instance')
        
        if ID_list is not None:
            self.individuals = [other.individuals[other.get_index_from_ID(id_)] for id_ in ID_list]
        
        elif index_list is not None:
            self.individuals = [other.individuals[idx] for idx in index_list]
            
        else:
            self.individuals = other.individuals
        
        if deepcopy:
            self.individuals = [copy.deepcopy(indi) for indi in self.individuals]

        self.update_id_list()
        self._sorted = other._sorted
        self._updated_crowding_distance = other._updated_crowding_distance
        self._updated_pareto_rank = other._updated_pareto_rank
        self._is_valid_database = other._is_valid_database
    
    def update_id_list(self) -> None:
        '''
        Update the list of IDs in the order of individuals.
        '''
        self._id_list = [indi.ID for indi in self.individuals]
        
    def sort_database(self, sort_type: int = 0) -> None:
        '''
        Sort the database by a certain type.
        
        Parameters:
        -----------
        sort_type: int
            Sort type defined in SettingsProblem.sort_type_dict.
            - 0: default, by dominance, crowding distance and potential
            - 1: sorting ID
            - 2: sorting x
            - 3: sorting y
            - 4: sorting objectives
            - 5: sorting type-2 output
            - 6: sorting crowding metrics
        '''
        for indi in self.individuals:
            indi.sort_type = sort_type
            
        if sort_type == 0:
            if not self.updated_pareto_rank or not self.updated_crowding_distance:
                raise ValueError('Pareto rank and crowding distance must be updated before sorting (type=0)')
        elif sort_type == 6:
            if not self.updated_crowding_distance:
                raise ValueError('Crowding distance must be updated before sorting (type=6)')
            
        self.individuals.sort()
        self._sorted = True
        self.update_id_list()
    
    #* Access data of individuals
    
    def get_index_from_ID(self, ID: int) -> int:
        '''
        Get index of an individual from its ID.
        '''
        return self._id_list.index(ID)
    
    def get_ID_from_index(self, index: int) -> int:
        '''
        Get ID of an individual from its index.
        '''
        return self.individuals[index].ID
    
    def get_largest_ID(self) -> int:
        '''
        Get the largest ID in the database.
        '''
        if self.size <= 0:
            return 0
        return np.max(self._id_list)
    
    def get_xs(self, scale: bool = False,
                ID_list: List[int] = None, 
                index_list: List[int] = None) -> np.ndarray:
        '''
        Get input variables of individuals in the database.
        
        Parameters:
        -----------
        scale: bool
            If True, return scaled input variables.
        ID_list: List[int]
            List of IDs of individuals to be selected. This has higher priority than `index_list`.
        index_list: List[int]
            List of index of individuals to be selected.
            
        Returns:
        --------
        xs: np.ndarray [n, n_input]
            Input variables of individuals in the database.
        '''
        if self.size <= 0:
            return None
    
        if ID_list is not None:
            nn = len(ID_list)
            xs = np.zeros([nn, self.problem.n_input])
            for i in range(nn):
                ii = self.get_index_from_ID(int(ID_list[i]))
                xs[i,:] = self.individuals[ii].x

        elif index_list is not None:
            nn = len(index_list)
            xs = np.zeros([nn, self.problem.n_input])
            for i in range(nn):
                xs[i,:] = self.individuals[index_list[i]].x

        else:
            nn = self.size
            xs = np.zeros([nn, self.problem.n_input])
            for i in range(nn):
                xs[i,:] = self.individuals[i].x

        if scale:
            xs = self.problem.scale_x(xs)

        return xs
    
    def get_ys(self, scale: bool = False,
                type_list: List[int] = None,
                ID_list: List[int] = None,
                index_list: List[int] = None) -> np.ndarray:
        '''
        Get output variables of individuals in the database.
        
        Parameters:
        -----------
        scale: bool
            If True, return scaled output variables.
        type_list: List[int]
            List of types of output variables to be selected.
        ID_list: List[int]
            List of IDs of individuals to be selected. This has higher priority than `index_list`.
        index_list: List[int]
            List of index of individuals to be selected.
            
        Returns:
        --------
        ys: np.ndarray [n, n_output]
            Output variables of individuals in the database.
        '''
        if self.size <= 0:
            return None
    
        if ID_list is not None:
            nn = len(ID_list)
            ys = np.zeros([nn, self.problem.n_output])
            for i in range(nn):
                ii = self.get_index_from_ID(int(ID_list[i]))
                ys[i,:] = self.individuals[ii].y

        elif index_list is not None:
            nn = len(index_list)
            ys = np.zeros([nn, self.problem.n_output])
            for i in range(nn):
                ys[i,:] = self.individuals[index_list[i]].y

        else:
            nn = self.size
            ys = np.zeros([nn, self.problem.n_output])
            for i in range(nn):
                ys[i,:] = self.individuals[i].y
    
        if scale:
            ys = self.problem.scale_y(ys)
    
        if not type_list is None:
            ys = self.problem.get_output_by_type(ys, type_list)
    
        return ys
    
    def get_unified_objectives(self, scale: bool = False,
                ID_list: List[int] = None,
                index_list: List[int] = None) -> np.ndarray:
        '''
        Return objective matrix with unified minimization direction.
        
        Parameters:
        -----------
        scale: bool
            If True, return scaled objectives.
        ID_list: List[int]
            List of IDs of individuals to be selected. This has higher priority than `index_list`.
        index_list: List[int]
            List of index of individuals to be selected.
            
        Returns:
        --------
        ys: np.ndarray [nn, n_objective]
            (Scaled) objective matrix with unified minimization direction.
        '''
        ys = self.get_ys(scale=scale, type_list=[1,-1],
                        ID_list=ID_list, index_list=index_list)
        if ys is None:
            return np.empty((0, 0))

        i_obj = 0
        for out_type in self.problem.problem_settings.output_type:
            if abs(out_type) != 1:
                continue
            if out_type == -1:
                ys[:, i_obj] = -ys[:, i_obj]
            i_obj += 1
        
        return ys
    
    #* Individual-level manipulation
    
    def check_duplication(self, x: np.ndarray,
                    is_scaled_x: bool = False
                    ) -> Tuple[List[bool]|bool, List[int]|int]:
        '''
        Check if the individual is duplicated.
        
        Note that the duplication is defined as the scaled distance between two individuals is less than a threshold, i.e.,
        `self.critical_scaled_distance`.
        
        Parameters:
        -----------
        x: np.ndarray [n, n_input] or [n_input]
            Input variables of the individual.
        is_scaled_x: bool
            If True, the input x is already scaled.
            
        Returns:
        --------
        is_duplicated: List[bool] or bool
            List of booleans indicating if the individuals are duplicated.
        closest_index: List[int] or int
            List of indices of the closest individuals.
        '''
        is_multiple = (x.ndim > 1)
        if is_multiple:
            n = x.shape[0]
            is_duplicated = [False] * n
            closest_index = [None] * n
        else:
            n = 1
            is_duplicated = False
            closest_index = None
        
        if self.size<=0:
            return is_duplicated, closest_index
        
        if not is_scaled_x:
            x = self.problem.scale_x(x)
        
        scaled_distance_matrix = self.problem.calculate_scaled_distance(
            x, self.get_xs(scale=True),
            is_scaled_x=True)
        
        min_dis = np.min(scaled_distance_matrix, axis=1) # [n]
        closest_index = np.argmin(scaled_distance_matrix, axis=1) # [n]
    
        crit = self.critical_scaled_distance()
        if is_multiple:
            for i in range(n):
                is_duplicated[i] = (min_dis[i] < crit)
        else:
            if min_dis[0] < crit:
                is_duplicated = True
            closest_index = closest_index[0]

        return is_duplicated, closest_index

    def add_individual(self, indi: Individual,
                    check_duplication: bool = True,
                    check_bounds: bool = True,
                    deepcopy: bool = True,
                    print_warning_info: bool = True,
                    ) -> Tuple[bool, str]:
        '''
        Add an individual to the database.
        
        Only the individual ID may be modified,
        other attributes are not modified.
        
        Parameters:
        -----------
        indi: Individual
            Individual to be added.
        check_duplication: bool
            If True, check if the individual is duplicated.
        check_bounds: bool
            If True, check if the individual is out of bounds.
        deepcopy: bool
            If True, the individual is copied.
        print_warning_info: bool
            If True, print warning information of the individual.
        
        Returns:
        --------
        added: bool
            True if the individual is added, False otherwise.
        warning_info: str
            Warning information of the individual.
        '''
        # Check problem
        if indi.problem != self.problem:
            raise ValueError('Individual problem does not match database problem')
        
        # Check bounds
        if check_bounds:
            if not self.problem.check_bounds_x(indi.x):
                text = f'Failed to add individual (ID={indi.ID}): x out of bounds'
                if print_warning_info:
                    print(f'>>> {text}')
                return False, text
        
        if deepcopy:
            indi = copy.deepcopy(indi)
        
        # Check duplication
        if self.size > 0:
            is_duplicated, closest_index = self.check_duplication(indi.x)
            if is_duplicated and check_duplication:
                text = f'Failed to add individual (ID={indi.ID}): duplicated with ID {closest_index} in database'
                if print_warning_info:
                    print(f'>>> {text}')
                return False, text
        
        # Assign ID
        original_ID = indi.ID
        if original_ID is not None:
            if indi.ID in self._id_list:
                indi.ID = self.get_largest_ID() + 1
        else:
            indi.ID = self.get_largest_ID() + 1
            
        # Add individual to database
        self.individuals.append(indi)
        self._id_list.append(indi.ID)
        self._sorted = False
        self._updated_crowding_distance = False
        self._updated_pareto_rank = False
        self._is_valid_database = False
            
        text = f'Added individual (ID={indi.ID:3d}, original ID={original_ID})'
        return True, text
    
    def delete_individual(self, ID: int = None, index: int = -1) -> None:
        '''
        Delete an individual from the database.
        
        Parameters:
        -----------
        ID: int
            ID of the individual to be deleted.
        index: int
            Index of the individual to be deleted.
        '''
        if self.size <= 0:
            raise Exception('Can not delete individual from empty database')
        
        if ID is None and index == -1:
            self.individuals.pop()
            self._id_list.pop()

        elif ID in self._id_list:
            ii = self.get_index_from_ID(ID)
            self.individuals.pop(ii)
            self._id_list.pop(ii)

        elif index>=0 and index<self.size:
            self.individuals.pop(index)
            self._id_list.pop(index)
        
        else:
            raise Exception('ID or index not valid (size=%d)'%(self.size), ID, index)
        
        self._sorted = False
        self._updated_crowding_distance = False
    
    def shrink_database(self, remaining_size: int, 
                reserve_ratio: float = 0.3) -> None:
        '''
        Shrink database to `remaining_size` by deleting
        worst individuals (based on Pareto rank and crowding distance).
        
        Parameters:
        -----------
        remaining_size: int
            Remaining size to shrink to.
        reserve_ratio: float
            Reserve ratio for database shrinking, i.e.,
            ratio of individuals that are directly kept.
        '''
        if self.size < remaining_size:
            return

        n_pop = self.size
        n_direct = int(reserve_ratio * remaining_size)
        if n_direct > 0:
            ii_sub = np.random.choice(n_pop, size=n_direct, replace=False)
            id_direct = [self.individuals[i].ID for i in ii_sub]
        else:
            id_direct = []

        i = 1
        while self.size > remaining_size:
            _id = self.individuals[-i].ID
            if _id in id_direct:
                i += 1
            else:
                self.delete_individual(ID=_id)
    
    def eliminate_invalid_individuals(self) -> None:
        '''
        Eliminate invalid individuals from the database.
        
        Invalid individuals are defined as:
        - the individual has no valid evaluation.
        - the individual is out of bounds.
        - the individual has constraint violations.
        - the constraints are updated in this function.
        '''
        # Loop backwards to avoid index shifting issues
        for i in range(self.size-1, -1, -1):
            
            indi = self.individuals[i]
            
            if not indi.valid_evaluation:
                self.delete_individual(index=i)
                continue
            
            if not self.problem.check_bounds_x(indi.x):
                self.delete_individual(index=i)
                continue
            
            if indi.y is not None:
                indi.eval_constraints()
            
            if indi.sum_violation > 0.0:
                self.delete_individual(index=i)
                continue
            
        self.update_id_list()
        self._sorted = False
        self._updated_crowding_distance = False
        self._updated_pareto_rank = False
        self._is_valid_database = True
    
    #* Database-level manipulation
    
    def get_sub_database(self,
                ID_list: List[int] = None, 
                index_list: List[int] = None,
                deepcopy: bool = True) -> 'Database':
        '''
        Create a sub-database from the database.
        
        Parameters:
        -----------
        ID_list: List[int]
            List of IDs of individuals to be selected. This has higher priority than `index_list`.
        index_list: List[int]
            List of index of individuals to be selected.
        deepcopy: bool
            If True, the individuals are copied.
        
        Returns:
        --------
        sub_database: Database
            Sub-database created from the database.
        '''
        if ID_list is not None and index_list is not None:
            raise ValueError('Only one of ID_list and index_list should be provided')

        sub_database = Database(self.problem, database_type='sub-database')

        if ID_list is not None:
            sub_database.individuals = [self.individuals[self.get_index_from_ID(id_)] for id_ in ID_list]
        
        if index_list is not None:
            sub_database.individuals = [self.individuals[idx] for idx in index_list]
        
        if deepcopy:
            sub_database.individuals = [copy.deepcopy(indi) for indi in sub_database.individuals]
        
        sub_database.update_id_list()

        return sub_database

    def get_intersection_with_database(self,
                        other: 'Database',
                        deepcopy: bool = True) -> 'Database':
        '''
        Get the intersection of the database with another database.
        
        The intersection is defined as:
        - the `x` and `y` of individuals are the same in both databases.
        - the intersection is a new database.
        - the intersection database is picked from this database.
        
        Parameters:
        -----------
        other: Database
            Another database to get the intersection with.
        deepcopy: bool
            If True, the individuals are copied.
        '''
        if other.problem is not self.problem:
            raise ValueError('Databases must share the same problem instance')

        def _same_xy(a: Individual, b: Individual) -> bool:
            if not np.array_equal(a.x, b.x):
                return False
            if a.y is None and b.y is None:
                return True
            if a.y is None or b.y is None:
                return False
            return np.array_equal(a.y, b.y)

        intersection = Database(self.problem, database_type='intersection')
        for indi in self.individuals:
            if any(_same_xy(indi, o) for o in other.individuals):
                intersection.individuals.append(indi)
        
        if deepcopy:
            intersection.individuals = [copy.deepcopy(indi) for indi in intersection.individuals]
        
        intersection.update_id_list()

        return intersection

    def merge_with_database(self,
                        other: 'Database',
                        deepcopy: bool = True,
                        log_func: Callable = None) -> None:
        '''
        Merge the database with another database.
        
        The merge is defined as:
        - the individuals are merged into this database.
        - the duplicated individuals are not added.
        
        Parameters:
        -----------
        other: Database
            Another database to merge with.
        deepcopy: bool
            If True, the individuals are copied.
        log_func: Callable
            Function to log the merge information.
            If None, print warning information to screen.
        '''
        if other.problem is not self.problem:
            raise ValueError('Databases must share the same problem instance')

        if other.size <= 0:
            return

        if log_func is None:
            print_warning_info = True
        else:
            print_warning_info = False

        for indi in other.individuals:
            added, warning_info = self.add_individual(indi, deepcopy=deepcopy,
                                            print_warning_info=print_warning_info)
            if not added and log_func is not None:
                log_func(warning_info, level=2, prefix='  - ')

        self.update_id_list()
        self._sorted = False
        self._updated_crowding_distance = False
        self._updated_pareto_rank = False
        self._is_valid_database = False

    def create_database_of_sub_problem(self,
                    sub_problem: Problem) -> 'Database':
        '''
        Create a database of a sub-problem.
        
        The sub-problem is defined as:
        - its names of `x` and `y` are subset of the current problem.
        - the other parameters can be different from the current problem.
        
        Parameters:
        -----------
        sub_problem: Problem
            Sub-problem to create the database.
        '''
        # Check if the sub-problem is a subset of the current problem
        if not sub_problem.is_subset_of(self.problem):
            raise ValueError('Sub-problem is not a subset of the current problem')
        
        # Create a new database for the sub-problem
        sub_database = Database(sub_problem, database_type='sub-database')
        
        # Mapping of `x` and `y` to the sub-problem
        parent_in = self.problem.data_settings.name_input
        parent_out = self.problem.data_settings.name_output
        ix = [parent_in.index(n) for n in sub_problem.data_settings.name_input]
        iy = [parent_out.index(n) for n in sub_problem.data_settings.name_output]
        
        # Add individuals to the sub-database
        # Pick components by parent-problem order, arranged as sub-problem order
        for indi in self.individuals:
            
            indi = copy.deepcopy(indi)
            
            indi.problem = sub_problem
            indi.name_problem = sub_problem.name
            
            indi.x = indi.x[ix].copy()
            indi._scaled_x = sub_problem.scale_x(indi.x)
            
            if indi.y is not None:
                indi.y = indi.y[iy].copy()
                indi._scaled_y = sub_problem.scale_y(indi.y)
                
            indi.eval_constraints()
            
            sub_database.add_individual(indi, deepcopy=False, print_warning_info=False)
            
        sub_database.update_id_list()        
        sub_database.evaluate_constraints()
        
        return sub_database

    #* Input and output
    
    def output_database_json(self, fname: str):
        '''
        Output database to JSON file.
        '''
        database_data = {
            'database_type': self.database_type,
            'individuals': [indi.data for indi in self.individuals]
        }
        
        with open(fname, 'w') as f:
            json.dump(database_data, f, indent=4)

    def read_database_json(self, fname: str):
        '''
        Read database from JSON file.
        '''
        with open(fname, 'r') as f:
            database_data = json.load(f)
            
        self.database_type = database_data['database_type']
        all_individuals = database_data['individuals']
        
        self.individuals = []
        for indi_data in all_individuals:
            
            indi = Individual(self.problem, x=np.array(indi_data['x']))
            
            for key, value in indi_data.items():
                
                if key == 'y':
                    if value is not None:
                        value = np.array(value)
                elif key == 'constraint_violations':
                    if value is not None:
                        value = np.array(value)
                    
                indi.__setattr__(key, value)
                
            self.individuals.append(indi)
        
        self._id_list = [indi.ID for indi in self.individuals]
        self._sorted = False
        self._updated_crowding_distance = False
        self._updated_pareto_rank = False
        self._is_valid_database = False
    
    @staticmethod
    def _format_sheet_left_align_and_auto_width(ws) -> None:
        '''
        Left-align all cells and auto-fit column widths for a worksheet.
        '''
        max_col_width = {}
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell_value = '' if cell.value is None else str(cell.value)
                col_idx = cell.column
                width = len(cell_value)
                if col_idx not in max_col_width or width > max_col_width[col_idx]:
                    max_col_width[col_idx] = width
        
        for col_idx, width in max_col_width.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max(width + 2, 10), 80)

    def json_to_excel(self, json_fname: str, excel_fname: str, sheet_name: str = 'database') -> None:
        '''
        Convert database JSON file to an Excel file.
        
        Excel format:
        - First row is header.
        - First column is `ID`.
        - Second column is `generation`.
        - Input/output variable components are expanded into independent columns,
          ordered by problem settings.
        '''
        with open(json_fname, 'r', encoding='utf-8') as f:
            database_data = json.load(f)
        
        all_individuals = database_data.get('individuals', [])
        
        input_names = list(self.problem.data_settings.name_input)
        output_names = list(self.problem.data_settings.name_output)
        
        # Keep x/y in required order; remaining fields can be in arbitrary order.
        exclude_keys = {'ID', 'generation', 'x', 'y'}
        other_keys = []
        for indi_data in all_individuals:
            for key in indi_data.keys():
                if key not in exclude_keys and key not in other_keys:
                    other_keys.append(key)
        
        header = ['ID', 'generation'] + input_names + output_names + other_keys
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(header)
        
        n_input = len(input_names)
        n_output = len(output_names)
        
        for indi_data in all_individuals:
            x = indi_data.get('x', [None] * n_input)
            y = indi_data.get('y', [None] * n_output)
            
            if x is None:
                x = [None] * n_input
            if y is None:
                y = [None] * n_output
            
            if len(x) < n_input:
                x = list(x) + [None] * (n_input - len(x))
            if len(y) < n_output:
                y = list(y) + [None] * (n_output - len(y))
            
            row = [
                indi_data.get('ID'),
                indi_data.get('generation'),
            ] + list(x[:n_input]) + list(y[:n_output])
            
            for key in other_keys:
                value = indi_data.get(key)
                if isinstance(value, list):
                    value = json.dumps(value, ensure_ascii=False)
                row.append(value)
            
            ws.append(row)
        
        # Additional sheet: data_settings
        ws_data = wb.create_sheet(title='data_settings')
        ds = self.problem.data_settings
        ws_data.append(['field', 'value'])
        ws_data.append(['name', ds.name])
        ws_data.append(['name_input', json.dumps(list(ds.name_input), ensure_ascii=False)])
        ws_data.append(['input_low', json.dumps(ds.input_low.tolist(), ensure_ascii=False)])
        ws_data.append(['input_upp', json.dumps(ds.input_upp.tolist(), ensure_ascii=False)])
        ws_data.append(['input_precision', json.dumps(ds.input_precision.tolist(), ensure_ascii=False)])
        ws_data.append(['name_output', json.dumps(list(ds.name_output), ensure_ascii=False)])
        ws_data.append(['output_low', json.dumps(ds.output_low.tolist(), ensure_ascii=False)])
        ws_data.append(['output_upp', json.dumps(ds.output_upp.tolist(), ensure_ascii=False)])
        ws_data.append(['output_precision', json.dumps(ds.output_precision.tolist(), ensure_ascii=False)])
        ws_data.append(['critical_scaled_distance', ds.critical_scaled_distance])

        # Additional sheet: problem_settings
        ws_problem = wb.create_sheet(title='problem_settings')
        ps = self.problem.problem_settings
        ws_problem.append(['field', 'value'])
        ws_problem.append(['name', ps.name])
        ws_problem.append(['name_data_settings', ps.name_data_settings])
        ws_problem.append(['output_type', json.dumps(list(ps.output_type), ensure_ascii=False)])
        ws_problem.append(['constraint_strings', json.dumps(list(ps.constraint_strings), ensure_ascii=False)])
        ws_problem.append(['n_constraint_functions', len(ps.constraint_functions)])
        ws_problem.append(['n_constraint', ps.n_constraint])
        ws_problem.append(['n_objective', ps.n_objective])
        
        self._format_sheet_left_align_and_auto_width(ws_data)
        self._format_sheet_left_align_and_auto_width(ws_problem)
        
        wb.save(excel_fname)

    #* Evaluation
    
    def evaluate_individuals(self,
                    mp_evaluation: MultiProcessEvaluation = None,
                    user_func: Callable = None,
                    prefix_folder_name: str = None) -> None:
        '''
        Evaluate the individuals (`y`) in the database,
        constraints are also evaluated.
        
        Parameters:
        -----------
        mp_evaluation: MultiProcessEvaluation
            Multi-process evaluation object.
            If None, use serial evaluation.
        user_func: Callable
            User-defined function to evaluate the individuals.
            If None, use external evaluation script.
        prefix_folder_name: str
            Prefix of the folder name for external evaluation.
            If None, use individual's ID as the folder name.
        
        Example:
        ---------
        >>> def user_func(x: np.ndarray, **kwargs) -> Tuple[bool, np.ndarray]:
        >>>     return True, np.array([np.sum(x**2)])
        
        Returns:
        --------
        None
        '''
        if self.size <= 0:
            return None
        
        if prefix_folder_name is None:
            prefix_folder_name = ''

        # Define folder names for external evaluation.
        xs = np.zeros((self.size, self.problem.n_input))
        list_name = []
        for i, indi in enumerate(self.individuals):
            xs[i, :] = indi.x
            if indi.ID is None:
                raise ValueError('Individual ID is None')
            list_name.append(prefix_folder_name + str(indi.ID))

        # Evaluate individuals.
        if mp_evaluation is not None:
            # Use mpEvaluation for both user_func and external_run modes.
            if callable(user_func):
                mp_evaluation.func = user_func
                list_succeed, ys = mp_evaluation.evaluate(
                    xs, list_name=None)
            else:
                mp_evaluation.func = None
                list_succeed, ys = mp_evaluation.evaluate(
                    xs, list_name=list_name, prob=self.problem)
                
            if ys.shape != (self.size, self.problem.n_output):
                raise ValueError(f'Invalid ys shape: {ys.shape} != [{self.size}, {self.problem.n_output}]')
                
        else:
            # Use serial evaluation.
            ys = np.zeros((self.size, self.problem.n_output))
            list_succeed = [False for _ in range(self.size)]
            
            for i, indi in enumerate(self.individuals):
                
                succeed = False

                if callable(user_func):
                    try:
                        succeed, y = user_func(indi.x)
                    except Exception as e:
                        print(f'    [user_func] failed for ID={indi.ID}: {e}')
                else:
                    succeed, y = self.problem.external_run(list_name[i], indi.x)

                if not succeed:
                    y = np.zeros(self.problem.n_output, dtype=float)
                elif not isinstance(y, np.ndarray):
                    raise ValueError(f'Invalid y type for ID={indi.ID}: {type(y)}')

                list_succeed[i] = succeed
                ys[i, :] = y

        # Update individuals.
        for i, indi in enumerate(self.individuals):
            
            y = ys[i, :]

            if list_succeed[i]:
                indi.y = y.copy()
                indi._scaled_y = self.problem.scale_y(indi.y)
                indi.valid_evaluation = True
                indi.eval_constraints()
            else:
                
                indi.y = None
                indi._scaled_y = None
                indi.valid_evaluation = False
                indi.constraint_violations = None
                indi.sum_violation = np.inf

        return None
